#!/usr/bin/env python3
"""
Import the Accepted/Rejected applicant workbook into public.application_profiles.

Usage:
  SUPABASE_URL=https://xxxx.supabase.co \
  SUPABASE_SERVICE_KEY=sb_secret_... \
  COHORT_SLUG=morph-by-tnx-cohort-2 \
  python3 scripts/import_application_profiles.py "/path/to/Accepted_Rejected List.xlsx"

- Matches applicants to participants by lower(email).
- Upserts on lower(email) so re-running is safe.
- Read-only data: only this script writes to the table (service role).
"""
import os
import sys
import json
import urllib.request

import openpyxl

# Map workbook column header -> table column. Headers are matched case-insensitively and
# trimmed. Unmapped columns are ignored.
HEADER_MAP = {
    "first name": "first_name",
    "last name": "last_name",
    "email address": "email",
    "phone number (whatsapp preferred)": "phone",
    "gender": "gender",
    "age range": "age_range",
    "institution / organization": "institution",
    "level of study": "level_of_study",
    "how would you describe your background": "background",
    "have you ever built or worked on a digital product before": "built_product_before",
    "what best describes your interest in this bootcamp?": "bootcamp_interest",
    "do you currently have a n idea you'd like to turn into a product?": "has_idea",
    "describe the idea you want to work on during the bootcamp": "idea_description",
    "who do you think would benetif the most from your solution?": "beneficiary",
    "what skills or knowledge are u most hoping to gain from this program?": "skills_hoping_to_gain",
    "in what ways would you be interested in contributing after the program?": "contribution_interest",
    "are you interested in being considered as a community lead  for this cohort?": "interested_community_lead",
    "why are you interested in supporting community coordination and peer engagement?": "community_lead_reason",
    "do you have any prior volunteering, leadership, or community management experience?": "prior_experience",
    "if yes, briefly describe your experience and your role": "prior_experience_detail",
    "what strengths would you bring to a learning community like this?": "community_strengths",
    "enter your scholarship code": "scholarship_code",
    "is there anything else you'd like us to know about you?": "anything_else",
    "how did you hear about us?": "heard_about_us",
    "do you have access to laptop/desktop computer?": "has_laptop",
    "do you have stable internet connection?": "has_internet",
    "what do you think makes you a good for this bootcamp?": "good_fit_reason",
}

# The long "giving back" prompt + commitment column vary in exact wording; match by prefix.
def map_header(raw: str) -> str | None:
    if not raw:
        return None
    key = " ".join(str(raw).strip().lower().split())
    if key in HEADER_MAP:
        return HEADER_MAP[key]
    if key.startswith("tnx solve focuses") or "why is giving back" in key:
        return "giving_back_importance"
    if key.startswith("the program requires weekly"):
        return "can_commit"
    return None


def clean(value):
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def supabase_request(url, key, method, path, body=None):
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(f"{url}/rest/v1/{path}", data=data, method=method)
    req.add_header("apikey", key)
    req.add_header("Authorization", f"Bearer {key}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Prefer", "resolution=merge-duplicates,return=representation")
    with urllib.request.urlopen(req) as resp:
        raw = resp.read().decode()
        return json.loads(raw) if raw else []


def main():
    if len(sys.argv) < 2:
        print("Provide the xlsx path as the first argument.")
        sys.exit(1)
    path = sys.argv[1]
    url = os.environ["SUPABASE_URL"].rstrip("/")
    key = os.environ["SUPABASE_SERVICE_KEY"]
    cohort_slug = os.environ.get("COHORT_SLUG", "")

    seed_participants = os.environ.get("SEED_PARTICIPANTS") == "1"
    seed_sheet = os.environ.get("SEED_SHEET", "Accepted")

    cohort_id = None
    if cohort_slug:
        cohorts = supabase_request(url, key, "GET", f"cohorts?slug=eq.{cohort_slug}&select=id")
        if cohorts:
            cohort_id = cohorts[0]["id"]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Optionally seed participants from the accepted sheet so profiles link by email.
    if seed_participants and cohort_id:
        seen = set()
        participants_payload = []
        for ws in wb.worksheets:
            if ws.title.strip().lower() != seed_sheet.strip().lower():
                continue
            header = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    if row and any(row):
                        header = [map_header(c) for c in row]
                    continue
                if not row or not any(row):
                    continue
                rec = {}
                for ci, ck in enumerate(header):
                    if ck and ci < len(row):
                        rec[ck] = clean(row[ci])
                email = (rec.get("email") or "").strip().lower()
                if not email or email in seen:
                    continue
                seen.add(email)
                name = " ".join(filter(None, [rec.get("first_name"), rec.get("last_name")])).strip()
                participants_payload.append({
                    "cohort_id": cohort_id,
                    "full_name": name or email,
                    "email": email,
                    "whatsapp": rec.get("phone"),
                    "source": rec.get("heard_about_us"),
                })
        for i in range(0, len(participants_payload), 50):
            batch = participants_payload[i : i + 50]
            supabase_request(url, key, "POST", "participants", batch)
        print(f"Seeded {len(participants_payload)} participants from '{seed_sheet}'.")

    # Build email -> participant id map for linking profiles.
    participant_by_email = {}
    if cohort_id:
        parts = supabase_request(url, key, "GET", f"participants?cohort_id=eq.{cohort_id}&select=id,email")
        for p in parts:
            if p.get("email"):
                participant_by_email[p["email"].strip().lower()] = p["id"]

    # Which sheets to import profiles from (comma-separated). Defaults to Accepted only.
    profile_sheets = [s.strip().lower() for s in os.environ.get("PROFILE_SHEETS", "Accepted").split(",")]

    rows_to_upsert = []
    for ws in wb.worksheets:
        if ws.title.strip().lower() not in profile_sheets:
            continue
        decision = ws.title.strip()
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                if row and any(row):
                    header = [map_header(c) for c in row]
                continue
            if not row or not any(row):
                continue
            record = {"decision": decision, "cohort_id": cohort_id}
            ts = None
            for col_index, col_key in enumerate(header):
                if col_index >= len(row):
                    continue
                val = row[col_index]
                if col_index == 0 and val is not None:
                    # first column is the form timestamp
                    try:
                        ts = val.isoformat()
                    except AttributeError:
                        ts = None
                if col_key:
                    record[col_key] = clean(val)
            record["submitted_at"] = ts
            email = (record.get("email") or "").strip().lower()
            if not email:
                continue
            record["email"] = email
            record["participant_id"] = participant_by_email.get(email)
            rows_to_upsert.append(record)

    print(f"Prepared {len(rows_to_upsert)} application rows.")
    matched = sum(1 for r in rows_to_upsert if r.get("participant_id"))
    print(f"Matched to participants: {matched}")

    # Upsert in batches on the email unique index.
    for i in range(0, len(rows_to_upsert), 50):
        batch = rows_to_upsert[i : i + 50]
        supabase_request(url, key, "POST", "application_profiles?on_conflict=email", batch)
        print(f"Upserted {min(i + 50, len(rows_to_upsert))}/{len(rows_to_upsert)}")

    print("Done.")


if __name__ == "__main__":
    main()
